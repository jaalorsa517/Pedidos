# -*- coding:utf-8 -*-
''' 
Modulo que crea la interfaz gráfica GUI
'''

import datetime
import os
import tkinter as tk
import tkcalendar as cld
import tkinter.filedialog as fdl
import tkinter.messagebox as msg
from modules import orderall, getmedida
from openpyxl import Workbook, load_workbook

hoy = datetime.date.today()
path = None
fecha = None

#Evento del boton archivo
def on_btArchivo_clic():
    global path
    path = fdl.askopenfilename()

    if ('/' in path):
        _t = path.split('/')
    elif ('\\' in path):
        _t = path.split('\\')

    _txt = _t[-1]
    btArchivo.configure(text=_txt)


#Evento del boton procesar
def on_btProcesar_clic():
    global path
    global fecha

    fecha = calendar.parse_date(calendar.get_date())

    if (path is not None and fecha is not None):

        if ('.xlsx' in path):

            pedido_path = 'Pedido.xlsx'

            if (os.path.isfile(pedido_path)):
                os.remove(pedido_path)

            pedido = Workbook()
            pedido.save(pedido_path)

            orderall.order(path, pedido_path, fecha)
            getmedida.medida(pedido_path, True)

            pedido = load_workbook(pedido_path)
            pedido.remove(pedido.worksheets[0])
            pedido.save(pedido_path)

            msg.showinfo('Finalizó el proceso', 'El proceso terminó con éxito')

        else:

            msg.showwarning('Archivo erróneo',
                            'Por favor seleccione un archivo XLSX')

    else:
        msg.showwarning('Falta seleccionar la fecha o el archivo',
                        'Por favor seleccione una fecha y/o un archivo')


#Ventana principal
win = tk.Tk()
win.resizable(0, 0)
win.title("Pedido")
win.geometry('400x350')

#Frame principal
f_main = tk.Frame(win)
f_main.pack(expand=True,fill=tk.BOTH)

f_g1 = tk.Frame(f_main)
f_g1.pack(expand=True)

f_g2 = tk.Frame(f_main)
f_g2.pack(expand=True)

f_g3 = tk.Frame(f_main)
f_g3.pack(expand=True)

#Etiqueta
label = tk.Label(f_g1, text='Archivo')
label.configure(font=('Arial',18)) #Configuracion para windows
label.grid(row=0, column=0)

#Boton ubicar XLSX
btArchivo = tk.Button(f_g1, text='Seleccionar...', command=on_btArchivo_clic)
btArchivo.config(font=('Arial',16)) #Configuracion para windows
btArchivo.grid(row=0, column=1)

#Calendario
calendar = cld.Calendar(f_g2,
                        selectmode='day',
                        year=hoy.year,
                        month=hoy.month,
                        day=hoy.day,
                        font=('Arial',16), #Configuracion para windows
                        locale='es_CO',
                        bordercolor='black',
                        background='#009600',
                        headersbackground ='#00b400',
                        normalbackground='#b4ffb4',
                        weekendbackground='#87ff87',
                        othermonthbackground='#32ff32',
                        othermonthwebackground='#32ff32',
                        othermonthweforeground ='black',
                        othermonthforeground ='black',
                        weekendforeground='black',
                        normalforeground='black')
calendar.pack(expand=True)

#Boton procesar
btProcesar = tk.Button(f_g3, text='Procesar', command=on_btProcesar_clic)
btProcesar.config(font=('Arial',16)) #Configuracion para windows
btProcesar.pack()

if __name__ == "__main__":
    win.mainloop()
