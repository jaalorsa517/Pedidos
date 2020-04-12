# -*- coding: utf-8 -*-
'''
Modulo que obtiene las medidas de los productos
'''

import click
from openpyxl import load_workbook
import modules.utility as ut


def medida(xlsx, sheet=False):
    _workbook = load_workbook(xlsx)
    _sheet = None

    if (not sheet):
        _sheet = ut.getSheet(_workbook)
    else:
        _sheet = _workbook.get_sheet_by_name(_workbook.sheetnames[-1])

    #Obtener Dic {row:producto}
    _list_productos = ut.getPedido(_sheet)

    for key, product in _list_productos.items():
        _sheet.cell(row=key, column=3, value=ut.get_unidad(product))
        ut.formato(_sheet['C{}'.format(key)], 'Arial', 14, False)
        ut.bordes(_sheet['C{}'.format(key)])

    _workbook.save(xlsx)
