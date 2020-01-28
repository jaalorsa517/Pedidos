# -*- coding: utf-8 -*-

import click
from openpyxl import load_workbook
import modules.utility as ut


def resume(xlsx):
    _workbook = load_workbook(xlsx)
    _sheet = ut.getSheet(_workbook)

    #Obtener Dic {row:producto}
    _list_productos = ut.getPedido(_sheet)

    #Obtener conjunto
    _con_produtos = sorted(set(_list_productos.values()))

    #Escribir titulo y encabezados
    row_max = sorted(_list_productos)[-1]
    cont = row_max + 2
    _sheet.cell(row=cont, column=1, value='RESUMEN')
    _sheet.merge_cells('A{col}:C{col}'.format(col=cont))
    ut.formato(_sheet['A{}'.format(cont)], 'Arial', 18, True)
    cont += 2

    _sheet.cell(row=cont, column=1, value='Nombre')
    ut.formato(_sheet['A{}'.format(cont)], 'Arial', 16, True)
    ut.bordes(_sheet['A{}'.format(cont)])

    _sheet.cell(row=cont, column=2, value='Cantidad')
    ut.formato(_sheet['B{}'.format(cont)], 'Arial', 16, True)
    ut.bordes(_sheet['B{}'.format(cont)])

    _sheet.cell(row=cont, column=3, value='Unidad')
    ut.formato(_sheet['C{}'.format(cont)], 'Arial', 16, True)
    ut.bordes(_sheet['C{}'.format(cont)])

    #Escribir el resumen
    for product in _con_produtos:
        cont += 1
        _sheet.cell(row=cont, column=1, value=product)
        ut.formato(_sheet['A{}'.format(cont)], 'Arial', 14, False)
        ut.bordes(_sheet['A{}'.format(cont)])

        _sheet.cell(row=cont,
                    column=2,
                    value='=SUMIF(A5:A{rm},A{c},B5:B{rm})'.format(rm=row_max,
                                                                  c=cont))
        ut.formato(_sheet['B{}'.format(cont)], 'Arial', 14, False)
        ut.bordes(_sheet['B{}'.format(cont)])

        _sheet.cell(row=cont, column=3, value=ut.get_unidad(product))
        ut.formato(_sheet['C{}'.format(cont)], 'Arial', 14, False)
        ut.bordes(_sheet['C{}'.format(cont)])

    _workbook.save(xlsx)
