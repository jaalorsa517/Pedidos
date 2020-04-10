# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import modules.utility as ut
import datetime

mes_dict = {
    'ene': 1,
    'feb': 2,
    'mar': 3,
    'abr': 4,
    'may': 5,
    'jun': 6,
    'jul': 7,
    'ago': 8,
    'sep': 9,
    'oct': 10,
    'nov': 11,
    'dic': 12
}


def dateToString(date_):
    for key, val in mes_dict.items():
        if (val is date_.month):
            _fecha = '{}{}'.format(date_.day, key)
            return _fecha
    return ''


def order(inventario, pedido, fecha):
    _workbook = load_workbook(inventario)
    _fecha = dateToString(fecha)
    if (len(_fecha) == 4):
        _fecha = '0{}'.format(_fecha)
    # _fechaDate = fecha

    inv = []

    for hoja in _workbook.sheetnames:
        a1 = _workbook[hoja][1][0]
        cliente = {}

        if (a1.comment is not None):
            cliente = ut.datosInComment(a1.comment.text)
        else:
            print("No existe comentario en la hoja {}.".format(hoja))

        if (a1.value is not None):
            cliente['neg'] = a1.value
        else:
            cliente['neg'] = hoja

        'Recorrer las primera fila'
        for fila in _workbook[hoja].iter_rows(min_col=2, max_row=1):
            for cell in fila:
                if (cell.value is not None):

                    cell_value = ''
                    sw = False

                    if (type(cell.value) == datetime.datetime):
                        cell_value = dateToString(cell.value)
                        if (len(cell_value) == 4):
                            cell_value = '0{}'.format(cell_value)

                    elif (type(cell.value) == str):
                        if ('/' in cell.value):
                            _cell_values = cell.value.split('/')
                            for i in _cell_values:
                                if (not (i.isnumeric())):
                                    sw = False
                                    break
                                else:
                                    sw = True

                        if (cell.value[0:2].isnumeric() and len(cell.value) > 5
                                and not sw):
                            cell_value = cell.value[0:5]

                        elif (cell.value[0].isnumeric() and len(cell.value) > 5
                              and not sw):
                            cell_value = '0{}'.format(cell.value[0:4])

                        elif (cell.value[0:2].isnumeric()
                              and len(cell.value) <= 5):
                            cell_value = cell.value

                        elif (cell.value[0].isnumeric()
                              and len(cell.value) == 4):
                            cell_value = '0{}'.format(cell.value)
                        elif (sw):
                            cell_value = dateToString(
                                datetime.datetime(day=int(_cell_values[0]),
                                                  month=int(_cell_values[1]),
                                                  year=int(_cell_values[2])))

                    if (_fecha.upper() == cell_value.upper()):
                        cliente['pedido'] = ut.getDataColumn(_workbook[hoja],
                                                             ini_row=3,
                                                             col=cell.column +
                                                             1)

        if ('pedido' in cliente):
            if (len(cliente['pedido']) > 0):
                #[{'nom':'','pedido':{}}]
                inv.append(cliente)

    #Se carga el xlsx pedido
    _workbook = load_workbook(pedido)
    #Se crea la nueva hoja
    _sheet = _workbook.create_sheet(_fecha)

    #Formato de la cabecera
    _sheet.merge_cells('A1:C1')
    _sheet['A1'].value = 'ANDES {}'.format(_fecha)
    _sheet['A1'].font = Font(name='Arial', size=20, bold=True)
    _sheet['A1'].alignment = Alignment(horizontal='center')

    _row = 2

    for i in inv:
        #if ('nom' in i and 'id' in i and 'tel' in i and 'email' in i):
        if ('nom' in i):
            #ut.cabecera(_sheet,_row,i['neg'],i['nom'],i['id'],i['tel'],i['email'])
            ut.cabecera(_sheet, _row, i['neg'], i['nom'])
        else:
            ut.cabecera(_sheet, _row, i['neg'])

        _row += 3
        ut.setPedido(_sheet, _row, i['pedido'])
        _row += len(i['pedido']) + 1

    _workbook.save(pedido)
