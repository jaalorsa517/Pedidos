# -*- encoding: utf-8 -*-

import click
from openpyxl.styles import Font, Border, Side


def getSheet(book):
    '''
    Función que devuelve una hoja de calculo seleccionada por el usuario.
    @param book: Libro xlsx
    @return: Hoja de calculo
    '''
    #INICIO DEL MENU
    op = ''
    i = 0
    for sheet in book.get_sheet_names():
        i += 1
        op += '{}.{}\n'.format(str(i), sheet)

    while True:
        try:
            r = click.prompt('Ingrese la opción deseada:\n{}'.format(op))
            sheet = book[book.get_sheet_names()[int(r) - 1]]
            break
        except ValueError:
            continue
        except IndexError:
            continue
    #FIN DEL MENU
    return sheet


def getPedido(sheet):
    '''
    Función que obtiene todos los pedidos a un diccionario
    @param sheet: Hoja de calculo.
    @return dic: Diccionario {row:string}.
    '''
    list = {}
    #Ciclo que recorre desde la fila ini_row y la columna col
    for col in sheet.iter_cols(min_row=1, max_col=1):
        for cell in col:
            #Evalua si contiene dato
            if (not (cell.value == None)):
                #Evalua si el contendio esta en negrita (encabezado)
                if (not cell.font.b):
                    list[cell.row] = cell.value
    return list


def getDataColumn(sheet, ini_row=1, col=1):
    '''
    Función que obtiene todos los pedidos a un diccionario
    @param sheet: Hoja de calculo.
    @paramDefault ini_row: Primera fila de la iteracion.
    @paramDefault col: Columna a la cual iterar.
    @return dic: Diccionario {string:string}.
    '''
    list = {}
    #Ciclo que recorre desde la fila ini_row y la columna col
    for colu in sheet.iter_cols(min_row=ini_row, max_col=col, min_col=col):
        for cell in colu:
            #Evalua si finalizó la tabla
            if (sheet[cell.row][0].value is None):
                break
            #Evalua si contiene dato
            if (not (cell.value is None) and (cell.value != 0)):
                list[sheet[cell.row][0].value] = cell.value
    return list


def datosInComment(coment):
    '''
    Funcion que separa un texto con el patron'\\n' y a
    su vez se vuleve a separar con el patron ':'
    _datos=coment.split('\\n')
    @param coment: String con el patron 
                    cc:123\\nnom:nom\\ntel:123\\email:email
    @return dic{cc:,nom:,tel:,email}
    '''
    _datos = coment.split('\n')
    datos_dic = {}
    for d in _datos:
        try:
            i = d.split(':')
            datos_dic[i[0]] = i[1]
        except Exception:
            continue
    return datos_dic


def cabecera(sheet, row, neg, nom='', id='', tel='', email=''):
    '''
    Realiza la cabecera del pedido
    '''
    n = 1
    _cabecera = ('Producto', 'Cantidad', 'Medida')
    sheet[row][0].value = 'Negocio: {}'.format(neg)
    sheet[row + n][0].value = 'Nombre: {}'.format(nom)
    # sheet[row + 2][0].value = 'Identifcacion: {}'.format(id)
    # sheet[row + 3][0].value = 'Telefono: {}'.format(tel)
    # sheet[row + 4][0].value = 'email: {}'.format(email)
    for j in range(2):
        formato(sheet[row + j][0], 'Arial', 18, True)
        sheet.merge_cells('A{col}:C{col}'.format(col=row + j))
    i = 0
    n = 2
    for c in _cabecera:
        sheet[row + n][i].value = c
        formato(sheet[row + n][i], 'Arial', 16, True)
        bordes(sheet[row + n][i])
        i += 1


def setPedido(sheet, row, pedido):
    '''
    Copia el pedido a la hoja
    '''
    _row = row
    for pro, can in pedido.items():

        sheet[_row][0].value = pro
        formato(sheet[_row][0], 'Arial', 14, False)
        bordes(sheet[_row][0])

        sheet[_row][1].value = can
        formato(sheet[_row][1], 'Arial', 14, False)
        bordes(sheet[_row][1])
        _row += 1


def formato(cell, font, size, b):
    '''
    Asigna la fuente a la celda
    '''
    cell.font = Font(name=font, sz=size, bold=b)


def bordes(celda):
    '''
    Asigna los bordes a la celda
    '''
    s = Side(border_style='hair', color='FF000000')
    celda.border = Border(left=s, right=s, top=s, bottom=s)


def get_unidad(product):
    '''
    Función encargada de poner la unidad de medida de productos.
    @param product: Nombre del producto
    @return: String con la medida correspondiente
    '''
    PAQUETE = ('Avena*6', 'Frescolanta*6', 'Kumis*6', 'Tampico*6',
               'Tampico*125*30', 'Yagur*6')
    SIXPACK = ('Leche deslactosada*1100ml*6', 'Leche entera*1100ml*6',
               'Leche montefrio*900*6', 'Leche prolinco*900*6',
               'Leche prolinco deslactosada*6', 'Leche semidescremada*1100*6')
    CAJA = ('Leche Ricura CAJA', 'Leche polvo prolinco*780 CAJA')
    RISTRA = ('Leche en polvo RISTRA', 'Kipe RISTRA','Leche condensada RISTRA')
    KILO = ('Queso o cuajada de segunda*kilo')

    for p in PAQUETE:
        if (p == product):
            return 'Paquete'

    for p in SIXPACK:
        if (p == product):
            return 'Sixpack'

    for p in CAJA:
        if (p == product):
            return 'Caja'

    for p in RISTRA:
        if (p == product):
            return 'Ristra'

    for p in KILO:
        if (p == product):
            return 'Kilo'

    return 'Unidad'
