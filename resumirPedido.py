# -*- coding: utf-8 -*-
import click
from openpyxl import load_workbook
import organizarPedido as org

@click.group()
def main():
    """Programa que organizará los pedidos"""
    pass

#METODO
@main.command()
@click.argument('xlsx')
def getResume(xlsx):
    """
    Resumirá los pedidos.
    Se requiere un archivo de excel, ocupando las 3 primeras columnas con la información 
    de la siguiente estructura:
    Nombre-Cantidad-Unidad.
    Donde Nombre es un String, Cantidad un float de 1 punto y Unidad es un String
    """
    _workbook = load_workbook(xlsx)
    _sheet=getSheet(_workbook)

    #Obtener Dic {row:producto}
    _list_productos = org.getProductos(_sheet)

    #Obtener conjunto
    _con_produtos = sorted(set(_list_productos.values()))

    #Escribir titulo y encabezados
    row_max = sorted(_list_productos)[-1]
    cont = row_max + 2
    _sheet.cell(row=cont, column=1, value='RESUMEN')
    cont += 2
    _sheet.cell(row=cont, column=1, value='Nombre')
    _sheet.cell(row=cont, column=2, value='Cantidad')
    _sheet.cell(row=cont, column=3, value='Unidad')

    #Escribir el resumen
    for product in _con_produtos:
        cont += 1
        _sheet.cell(row=cont, column=1, value=product)
        _sheet.cell(row=cont,
                    column=2,
                    value='=SUMIF(A5:A{rm},A{c},B5:B{rm})'.format(rm=row_max,
                                                                  c=cont))
        _sheet.cell(row=cont, column=3, value=org.get_unidad(product))

    _workbook.save(xlsx)

@main.command()
@click.argument('xlsx')
def getMedida(xlsx):
    """
    Completará la columna de medida.
    Se requiere un archivo de excel, ocupando las 3 primeras columnas con la información 
    de la siguiente estructura:
    Nombre-Cantidad-Unidad.
    Donde Nombre es un String, Cantidad un float de 1 punto y Unidad es un String
    """
    _workbook = load_workbook(xlsx)

    _sheet=getSheet(_workbook)
    
    #Obtener Dic {row:producto}
    _list_productos = org.getProductos(_sheet)

    for key, product in _list_productos.items():
        _sheet.cells(row=key,column=3,value=org.get_unidad(product))
    
    _workbook.save(xlsx)


def getSheet(book):
    '''
    Función que devuelve una hoja de calculo seleccionada por el usuario.
    @param book: Libro xlsx
    @return: Hoja de calculo
    '''
    #INICIO DEL MENU
    op = ''
    i = 0
    for sheet in book.get_sheet_names():
        i += 1
        op += '{}.{}\n'.format(str(i), sheet)

    while True:
        try:
            r = click.prompt('Ingrese la opción deseada:\n{}'.format(op))
            sheet = book[book.get_sheet_names()[int(r) - 1]]
            break
        except ValueError:
            continue
        except IndexError:
            continue
    return sheet
    #FIN DEL MENU

    


#PUNTO DE ENTRADA
if __name__ == "__main__":
    main()